import re
from django.shortcuts import render
from django.http import HttpResponse
from django.views.generic import ListView
from users.models import UserData
import requests
from django.core.paginator import Paginator
from .forms import call_analyse
import openpyxl
import os


API_KEY = [] #Список API-ключей пользователя со страницы сравнения с Excel файлом

data_project_dict = {} #Словарь со всеми проектами пользователя на Roistat (ключ - номер проекта, значение - название проекта)

ALL_CALLS = [] #Список всех звонков проекта на Roistat

name_project = ['Обновите данные'] #Имя выбранного проекта

EXCEL_CALLS = [] #Список номеров телефонов из excel файла для проверки

UNIQ_NO_NUM_PHONE = [] #Список уникальных/повторяющихся телефонов из excel файла после проверки всех номеров проекта


def index(request):
	if request.method == 'POST':
		form = call_analyse(request.POST)  # подгружаем форму на страничку
		if request.POST.get('roistat_api_connect'):#если нажали кнопку отобразить данные с ройстат, то подргужаем данные по всем проектам указанного АПИ-ключа
			API_KEY.clear()  # чистим список с АПИ-ключами перед тем, как добавить новый ключ
			API_KEY.append(request.POST.get('roistat_api_connect'))#получаем API-ключ и добавляем его в список.
			CURRENT_ROISTAT_ALL_PROJECT = (f'https://cloud.roistat.com/api/v1/user/projects?key={API_KEY[0]}')#загружаем все проекты доступные по этому АПИ-ключу
			project = requests.get(CURRENT_ROISTAT_ALL_PROJECT)#получаем все проекты пользователя на Roistat по апи ключу
			data_project = project.json()#переводим в json
			if data_project['status'] == 'error':
				context = {'error_message_api': 'Ошибка аутентификации. Проверьте правильность введённых данных.'}
			else:
				data_project_dict.update({i['id']:i['name'] for i in data_project['projects']})#записываем все проекты в словарь (ключ - номер проекта, значение - название проекта)
				form = call_analyse(request.POST, request.FILES)#подгружаем форму
				context = {'projects':data_project_dict, 'form':form}

		elif request.POST.get('num_project'):
			NUM_PROJECT = request.POST.get('num_project')#получаем номер (ключ) проекта из словара
			try:
				CURRENT_ROISTAT_API_CALLTRACKING = (f'https://cloud.roistat.com/api/v1/project/calltracking/call/list?key={API_KEY[0]}&project={NUM_PROJECT}')  # получаем список всех звонков по данному проекту
				status = requests.post(CURRENT_ROISTAT_API_CALLTRACKING)#получаем все звонки по выбранному проекту
				data_status = status.json()#переводим в json
			except:
				data_status = {'data':[{'caller':'Обновите данные'}]}
			ALL_CALLS.clear()  # чистим список перед циклом, чтобы выводить значение только по одному выбранному проекту и не суммировать с другими
			name_project.append(data_project_dict.get(
				int(NUM_PROJECT)))  # получаем название (значение) проекта из словара и добавляем в список

			################Блок сортировки звонков по датам#####################
			DATE_START = request.POST.get('date_start')#Начальная дата для сортировки
			DATE_END = request.POST.get('date_end')#Дата конца для сортировки
			if DATE_START and DATE_END:# если две даты для сортировки введены пользователем
				for i in data_status['data']:
					if i['date'] >= DATE_START+'T00:00:00+0000' and i['date'] <= DATE_END+'T23:59:59+9999':
						ALL_CALLS.append(i['caller'])
			elif DATE_START:# если только дата начала сортировки введена пользователем
				for i in data_status['data']:
					if i['date'] >= DATE_START+'T00:00:00+0000':
						ALL_CALLS.append(i['caller'])
			elif DATE_END:# если только дата конца сортировки введена пользователем
				for i in data_status['data']:
					if i['date'] <= DATE_END+'T23:59:59+9999':
						ALL_CALLS.append(i['caller'])
			else:
				for i in data_status['data']:
					ALL_CALLS.append(i['caller'])  # добавляем в список всех caller'ов
			################Блок работы с excel файлом#################
			EXCEL_CALLS.clear()#Чистим список с номерами телефонов из excel файла перед циклом, чтобы выводить значения только из текущего файла и не сумировать с предыдущими
			UNIQ_NO_NUM_PHONE.clear()#Чистим список с номерами телефонов
			excel_file = request.FILES["excel_file"]#перехватываем файл
			ext = os.path.splitext(excel_file.name)[1]  # [0] returns path+filename
			valid_extensions = ['.xlsx', '.xls'] #допустимый формат файла
			if not ext.lower() in valid_extensions:#проверка формата файла
				context = {'projects': data_project_dict, 'calls': ALL_CALLS, 'name_of_project': name_project[-1],
						   'form': form, 'error_message':'Недопустимый формат файла.'}
				return render(request, 'web/index.html', context=context)

			CHOICE = request.POST.get('uniq_num')#выбор: нам нужны уникальные номера телефонов или повторяющиеся
			RANGE_CELLS_START = request.POST.get('range_cells_start') # начало диапазона ячеек с номерами телефонов для сравнения в excel файле
			RANGE_CELLS_END = request.POST.get('range_cells_end') # окончание диапазона ячеек с номерами телефонов для сравнения в excel файле
			RANGE_CELLS_START = RANGE_CELLS_START.upper()
			RANGE_CELLS_END = RANGE_CELLS_END.upper()
			try:
				wb = openpyxl.load_workbook(excel_file)# создаем книгу
				ws = wb.active# делаем единственный лист активным
				for cell in ws[RANGE_CELLS_START+':'+RANGE_CELLS_END]:#цикл по ячейкам указанного диапазона
					for value in cell:#цикл по значениям ячеек указанного диапазона
						if value.value: #отсеиваем пустые ячейки
							EXCEL_CALLS.append(str(value.value))#добавляем в список все значения ячеек указанного диапазона

			except:
				context = {'projects': data_project_dict, 'calls': ALL_CALLS, 'name_of_project': name_project[-1],
						   			   'form': form, 'error_message_range': 'Неверный ввод данных'}
				return render(request, 'web/index.html', context=context)

			if CHOICE == 'uniq': # уникальные номера телефонов для сравнения
				for num_phone in EXCEL_CALLS:
					if num_phone not in ALL_CALLS:
						UNIQ_NO_NUM_PHONE.append(num_phone)
			else:
				for num_phone in EXCEL_CALLS: # повторяющиеся номера телефонов для сравнения
					if num_phone in ALL_CALLS:
						UNIQ_NO_NUM_PHONE.append(num_phone)
			paginator = Paginator(UNIQ_NO_NUM_PHONE,50)  # пагинация результата по 50 номеров телефона на страницу
			page = request.GET.get('page', 1)
			calls_paginator = paginator.get_page(page)
			context = {'projects': data_project_dict, 'calls':ALL_CALLS, 'name_of_project':name_project[-1], 'paginator': calls_paginator,
					   'form':form, 'excel_save':'ready'}
		return render(request, 'web/index.html', context=context)
	if request.method == 'GET': # пагинация результата по 50 номеров телефона на страницу методом Гет.
		paginator = Paginator(UNIQ_NO_NUM_PHONE, 50)
		page = request.GET.get('page', 1)
		calls_paginator = paginator.get_page(page)
		form = call_analyse(request.POST, request.FILES)
		################Сохраняем результат работы в excel файле#################
		if request.GET.get('excel_save'):
			output = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet, application/vnd.ms-excel')#тип файла в котором сохраняем (соответственно, excel)
			output['Content-Disposition'] = 'attachment; filename=' + 'Result.xlsx' #название его Result.xlsx
			wb = openpyxl.Workbook()
			ws = wb.active
			ws.title = "Nums_Phone" #название листа, в котором будет сохраняться результат
			r=1
			for i in UNIQ_NO_NUM_PHONE:
				ws.cell(row=r, column=1).value = i #результат сохраняем в 1 столбике
				r += 1
			wb.save(output)
			return output
		context = {'projects': data_project_dict, 'name_of_project':name_project[-1], 'paginator': calls_paginator}
		return render(request, 'web/index.html', context=context)


# if not re.match(r'^[:\]\[A-Z0-9]+$', RANGE_CELLS):
# 	context = {'projects': data_project_dict, 'calls': ALL_CALLS, 'name_of_project': name_project[-1],
# 			   'paginator': calls_paginator, 'form': form,
# 			   'error_message_range': 'Используйте только латинские буквы. Запись должна быть в формате W4:W37'}
# 	return render(request, 'web/index.html', context=context)
# else:
# 	wb = openpyxl.load_workbook(excel_file)#читаем excel файл
# 	ws = wb.active#читаем
# 	for i in ws[RANGE_CELLS]:
# 		for j in i:
# 		   EXCEL_CALLS.append(str(j.value))
# 	for i in EXCEL_CALLS:
# 		if i not in ALL_CALLS:
# 			UNIQ_NUM_PHONE.append(i)
# 	print(UNIQ_NUM_PHONE)



# def index(request):
# 	if request.method == 'POST':
# 		if request.POST.get('roistat_api_connect'):
# 			API_KEY.append(request.POST.get('roistat_api_connect'))#получаем API-ключ и добавляем его в список.
# 			CURRENT_ROISTAT_ALL_PROJECT = (f'https://cloud.roistat.com/api/v1/user/projects?key={API_KEY[-1]}')
# 			project = requests.get(CURRENT_ROISTAT_ALL_PROJECT)#получаем все проекты пользователя на Roistat по апи ключу
# 			data_project = project.json()
# 			data_project_dict.update({i['id']:i['name'] for i in data_project['projects']})#записываем все проекты в словарь (ключ - номер проекта, значение - название проекта)
# 			context = {'projects':data_project_dict}
#
# 		elif request.POST.get('num_project'):
# 			NUM_PROJECT = request.POST.get('num_project')#получаем номер (ключ) проекта из словара
# 			CURRENT_ROISTAT_API_CALLTRACKING = (f'https://cloud.roistat.com/api/v1/project/calltracking/call/list?key={API_KEY[-1]}&project={NUM_PROJECT}')
# 			status = requests.post(CURRENT_ROISTAT_API_CALLTRACKING)#получаем список всех звонков по данному проекту
# 			data_status = status.json()
# 			ALL_CALLS.clear() #чистим список перед циклом, чтобы выводить значение только по одному выбранному проекту
# 			name_project.append(data_project_dict.get(int(NUM_PROJECT)))#получаем название (значение) проекта из словара и добавляем в список
# 			for i in data_status['data']:
# 				ALL_CALLS.append(i['caller'])#добавляем в список всех caller'ов
# 			paginator = Paginator(ALL_CALLS, 50)
# 			page = request.GET.get('page', 1)
# 			calls_paginator = paginator.get_page(page)
# 			context = {'projects': data_project_dict, 'calls':ALL_CALLS, 'name_of_project':name_project[-1], 'paginator': calls_paginator}
# 		return render(request, 'web/index.html', context=context)
# 	if request.method == 'GET':
# 		paginator = Paginator(ALL_CALLS, 50)
# 		page = request.GET.get('page', 1)
# 		calls_paginator = paginator.get_page(page)
# 		context = {'projects': data_project_dict,'calls':ALL_CALLS, 'name_of_project':name_project[-1], 'paginator': calls_paginator}
# 		return render(request, 'web/index.html', context=context)
# 	else:
# 		return render(request, 'web/index.html')





# def index(request):
# 	API_KEY = '194b851bf90e8e91735117613de5a4b8'
# 	CURRENT_ROISTAT_ALL_PROJECT = (f'https://cloud.roistat.com/api/v1/user/projects?key={API_KEY}')
# 	project = requests.get(CURRENT_ROISTAT_ALL_PROJECT)
# 	data_project = project.json()
# 	data_project_dict = {i['id']:i['name'] for i in data_project['projects']}
# 	context = {'context':data_project_dict}
# 	return render(request, 'web/index.html', context=context)


# def index(request):
# 	if request.method == 'POST':
# 		global API_KEY
# 		API_KEY = request.POST.get('roistat_api_connect')
# 		CURRENT_ROISTAT_ALL_PROJECT = (f'https://cloud.roistat.com/api/v1/user/projects?key={API_KEY}')
# 		project = requests.get(CURRENT_ROISTAT_ALL_PROJECT)
# 		data_project = project.json()
# 		global data_project_dict
# 		data_project_dict = {i['id']:i['name'] for i in data_project['projects']}
# 		context = {'projects':data_project_dict}
# 		return render(request, 'web/index.html', context=context)
# 	elif request.method == 'GET':
# 		NUM_PROJECT = request.GET.get('num_project')
# 		print(API_KEY)
# 		CURRENT_ROISTAT_API_CALLTRACKING = (f'https://cloud.roistat.com/api/v1/project/calltracking/call/list?key={API_KEY}&project={NUM_PROJECT}')
# 		status = requests.post(CURRENT_ROISTAT_API_CALLTRACKING)
# 		data_status = status.json()
# 		context = {'projects': data_project_dict, 'calls':data_status}
# 		return render(request, 'web/index.html', context=context)
# 	else:
# 		return render(request, 'web/index.html')
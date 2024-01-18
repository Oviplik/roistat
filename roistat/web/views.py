from django.shortcuts import render, redirect, HttpResponse
from django.views.generic import ListView
from users.models import UserData
import requests
from .forms import call_analyse
import openpyxl
import os
from datetime import date
import time
import json

def manual(request):
	return render(request, 'web/manual.html')

def index(request):
	data_project_dict = {}  # Словарь со всеми проектами пользователя на Roistat (ключ - номер проекта, значение - название проекта)

	if request.method == 'POST':
		API_KEY = request.POST.get('roistat_api_connect').encode("utf-8").decode("latin1")  # получаем API-ключ и добавляем его в список.
		headers = {
			'Content-type': 'application/json',
			'Api-key': f'{API_KEY}',
		}
		response = requests.get('https://cloud.roistat.com/api/v1/user/projects', headers=headers)#загружаем все проекты доступные по этому АПИ-ключу
		data_project = response.json()  # переводим в json
		if data_project['status'] == 'error':
			context = {'error_message_api': 'Ошибка аутентификации. Проверьте правильность введённых данных.'}
		else:
			data_project_dict.clear()
			data_project_dict.update({i['id']:i['name'] for i in data_project['projects']})#записываем все проекты в словарь (ключ - номер проекта, значение - название проекта)
			form = call_analyse(request.POST, request.FILES)#подгружаем форму
			context = {'projects':data_project_dict, 'form':form, 'api_key':API_KEY}
		return render(request, 'web/project.html', context=context)

	return render(request, 'web/index.html')


def project(request):

	ALL_CALLS = []  # Список всех звонков проекта на Roistat
	ALL_CALLS_DUPLICATE = []  # Список всех звонков проекта на Roistat дупликат для сортировки лидов по датам и оставления в списки только самой ранней заявки
	EXCEL_CALLS = []  # Список номеров телефонов из excel файла для проверки
	UNIQ_NO_NUM_PHONE = []  # Список пересекающихся телефонов из excel файла после проверки всех номеров проекта

	if request.method == 'POST':
		form = call_analyse(request.POST)  # подгружаем форму на страничку
		API_KEY = request.POST.get('roistat_api_connect') #получаем api-ключ и возвращаем на шаблон страницы (и так до бесконечности, пока обновляем страницу работы с проектами по данному апи ключу)
		NUM_PROJECT = request.POST.get('num_project')#получаем номер (ключ) проекта из словара
		DATE_START = request.POST.get('date_start')  # Начальная дата для сортировки
		DATE_END = request.POST.get('date_end')  # Дата конца для сортировки
		data_project_dict = request.POST.get('project_names') # получаем все проекты и их id в str
		data_project_dict = eval(data_project_dict) #переводим str в dict и возвращаем на шаблон страницы (и так до бесконечности , пока обновляем страницу работы с проектами по данному апи ключу)
		name_project = data_project_dict[int(NUM_PROJECT)]#имя проекта с которым работаем

		if DATE_START:
			DATE_START = str(DATE_START)
		else:
			DATE_START = '2014-01-01T00:00:00+0300'
		if DATE_END:
			DATE_END = str(DATE_END)
		else:
			DATE_END = str(date.today()) + 'T23:59:59+0300'

		try:
			# получаем список всех лидов по данному проекту из API Управления заявками без CRM
			headers = {
				'Content-type': 'application/json',
				'Api-key': f'{API_KEY}',
			}
			params = {
				'project': f'{NUM_PROJECT}',
			}
			json_data = {
				'period': {
					'from': f'{DATE_START}',
					'to': f'{DATE_END}',
				},
			}
			response = requests.post(
				'https://cloud.roistat.com/api/v1/project/analytics/list-orders',
				params=params,
				headers=headers,
				json=json_data,
			)
			data_status = response.json()#переводим в json

			ALL_CALLS.clear()  # чистим список перед циклом, чтобы выводить значение только по одному выбранному проекту и не суммировать с другими
			for i in data_status['orders']:
				ALL_CALLS.append([i['id'], i['client_phones'][0], i['creation_date']])  # добавляем в список списки всех caller'ов с id заявки и датой создания
				ALL_CALLS_DUPLICATE.append([i['id'], i['client_phones'][0], i['creation_date']])  # добавляем дупликаты в список списки всех caller'ов с id заявки и датой создания

			for i in ALL_CALLS:
				for j in ALL_CALLS_DUPLICATE:
					if i[1] == j[1] and i[2] > j[2]:
						ALL_CALLS.remove(i)
		except:
				data_status = {'data':[{'caller':'Обновите данные'}]}

		################Блок работы с excel файлом#################
		EXCEL_CALLS.clear()#Чистим список с номерами телефонов из excel файла перед циклом, чтобы выводить значения только из текущего файла и не сумировать с предыдущими
		UNIQ_NO_NUM_PHONE.clear()#Чистим список с номерами телефонов
		excel_file = request.FILES["excel_file"]#перехватываем файл
		ext = os.path.splitext(excel_file.name)[1]  # [0] returns path+filename
		valid_extensions = ['.xlsx', '.xls'] #допустимый формат файла
		if not ext.lower() in valid_extensions:#проверка формата файла
			context = {'projects': data_project_dict,  'api_key':API_KEY, 'form': form, 'error_message':'Недопустимый формат файла.'}
			return render(request, 'web/project.html', context=context)

		try:
			wb = openpyxl.load_workbook(excel_file)# создаем книгу
			ws = wb.active# делаем единственный лист активным
			for (a, b) in ws.iter_rows(max_col=2):
				try:
					if float(b.value)>0:
						EXCEL_CALLS.append([str(a.value),str(b.value)])
				except:
					pass
		except:
			context = {'projects': data_project_dict, 'api_key':API_KEY,
								   'form': form, 'error_message_range': 'Неверный ввод данных'}
			return render(request, 'web/project.html', context=context)
		for num_phone in ALL_CALLS: # повторяющиеся номера телефонов для сравнения
			for num_phone_excel in EXCEL_CALLS:
				if num_phone[1] in num_phone_excel[0]:
					UNIQ_NO_NUM_PHONE.append([num_phone[1],num_phone[0],num_phone_excel[1]])

		len_UNIQ_NO_NUM_PHONE = len(UNIQ_NO_NUM_PHONE)
		len_ALL_CALLS_DUPLICATE = len(ALL_CALLS_DUPLICATE)#Общее количество заявок проекта в разделе «Управление заявками без CRM»
		len_EXCEL_CALLS = len(EXCEL_CALLS)#Общее количество номеров телефонов с оплатой в эксель файле

		global UNIQ_NO_NUM_PHONE_in_excel
		UNIQ_NO_NUM_PHONE_in_excel = UNIQ_NO_NUM_PHONE
		global API_KEY_for_change
		API_KEY_for_change = API_KEY
		global NUM_PROJECT_for_change
		NUM_PROJECT_for_change = NUM_PROJECT

		context = {'projects': data_project_dict, 'name_project': name_project, 'len_list_order':len_UNIQ_NO_NUM_PHONE, 'len_common_list':len_ALL_CALLS_DUPLICATE, 'len_list_call':len_EXCEL_CALLS, 'api_key':API_KEY, 'form':form, 'result':'result'}
		return render(request, 'web/project.html', context=context)
	if request.GET.get('excel_save'):
		################Сохраняем результат работы в excel файле#################
		output = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet, application/vnd.ms-excel')#тип файла в котором сохраняем (соответственно, excel)
		output['Content-Disposition'] = 'attachment; filename=' + 'Result.xlsx' #название его Result.xlsx
		wb = openpyxl.Workbook()
		ws = wb.active
		ws.title = "Nums_Phone" #название листа, в котором будет сохраняться результат
		r=2
		ws.cell(row=1, column=1).value = 'phone'
		ws.cell(row=1, column=2).value = 'id_order'
		ws.cell(row=1, column=3).value = 'cost'
		for i in UNIQ_NO_NUM_PHONE_in_excel:
			ws.cell(row=r, column=1).value = i[0] #телефон лида сохраняем в 1 столбике
			ws.cell(row=r, column=2).value = i[1]  #id заказа сохраняем в 2 столбике
			ws.cell(row=r, column=3).value = i[2]  # cost заказа сохраняем в 3 столбике
			r += 1
		wb.save(output)
		return output
	if request.GET.get('roistat_change_orders'):
		try:
			for i in UNIQ_NO_NUM_PHONE_in_excel:
				headers = {
					'Content-type': 'application/json',
					'Api-key': f'{API_KEY_for_change}',
				}

				params = {
					'project': f'{NUM_PROJECT_for_change}',
				}
				json_data = {"id": i[1], "price": i[2], "status": "1", "paid_date": str(date.today())+"T00:00:00+0300"}
				data = json.dumps(json_data)
				data = data.encode()
				response = requests.post('https://cloud.roistat.com/api/v1/project/leads/lead/update', params=params, headers=headers, data=data)
				time.sleep(0.2)
			context = {'status_roistat_change':'Данные успешно обновлены на Roistat'}
			return render(request, 'web/project.html', context=context)
		except:
			context = {'status_roistat_change': 'Не удалось обновить данные на Roistat'}
			return render(request, 'web/project.html', context=context)
	return render(request, 'web/project.html')


